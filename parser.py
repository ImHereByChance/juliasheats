import openpyxl 
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string
from collections import namedtuple
import re


Fieldstuple = namedtuple('Fields', 'varieties custumers numbers pieces \
				    totals    prices    amounts codes  \
				    codes_singleUse  rates_singleUse\
				    codes_multiUse   deposits_multiUse, rents_multiUse')


def find_main_data(book, sheet:str):
	"""
	finds desposition of rows containing main columns such as 
	varieties, custumers, numbers, pieces, totals, prices, amounts, codes.
	Returns tuple(beginning_row, ending_row)
	"""
	page = book.get_sheet_by_name(sheet)
	data_beginning_row = None
	data_range = 0
	for cell in page['A']:
		if cell.value in ('date', 'Ship\ndate'):
			data_beginning_row = cell.row + 1 
			if isinstance(page[f'A{data_beginning_row}'].value, float):  # check of row follows next to data_beginng row 
				continue
			else:
				raise ValueError("Error: 'date'-row is found but float-type date-cell don't follow further")
		if data_beginning_row is not None and type(cell.value) == float:
			data_range += 1
		elif data_beginning_row is not None and data_range > 1:
			break
	if data_beginning_row is None:
		print(f'no main data finded on "{sheet}"')
		return
	return data_beginning_row, data_beginning_row + data_range - 1


def find_quantity_columns(book, sheet:str, mainData_range):
	"""
	Finds desposition of columns: numbers, pieces, totals, prices, amounts, codes
	Returns tuple containing letter of each column in order.
	"""
	page = book.get_sheet_by_name(sheet)  
	data_beginning_row = str(mainData_range[0])
	values_list = page[data_beginning_row]

	mask = {1: (int,), 2: (int,), 3: (int, float),
			4: (int, float), 5: (str,), 6: (int,)}
	pos = 1
	beginning = None
	ending = None
	for i in values_list:
		if type(i.value) in mask[pos]:
			if beginning is None:
				beginning = i.column
			pos +=1
			if pos == 6:
				ending = i.column+1
				break
		else:
			beginning = None
			pos = 1
	if beginning is None:
		raise ValueError(f"couldn't find 'Number' 'Pieces' 'Total' 'Price' 'Amount' columns layout on the {sheet}" )
	return tuple(get_column_letter(i) for i in range(beginning, ending+1))


def find_additional_section(book, sheet:str, section:str):
	"""
	Finds "Single use packaging" and "Multi use packaging" sections on the page. 
	Returns tuple of (beginning_row, ending_row) row-numbers.
	"""
	page = book.get_sheet_by_name(sheet)
	data_beginning_row = None
	data_range = 0
	c = 0
	for cell in page['A']:
		if cell.value == section: 
			if not page[f'A{cell.row+1}'].value == 'Date':  
				raise ValueError(f'error during searching {section}range: \
								 "Date" row don`t follow after "{section}" row')
			if not is_longFormat_date(page[f'A{cell.row+2}'].value):
				if is_gap_after_date(page[f'{cell.row+2}']):
					data_beginning_row = cell.row + 3
					continue
				else:
					raise ValueError(f'couldn`t define {section} on {page}')
			data_beginning_row = cell.row + 2
			continue
		if data_beginning_row is not None:
			if is_longFormat_date(cell.value):
				data_range += 1
				continue
			elif cell.value == 'Total':
				break
	if data_beginning_row is None:
		print(f'no "{section}" section finded on "{sheet}"')
		return
	return data_beginning_row, data_beginning_row + data_range - 1


def find_rates_singleUse(book, sheet:str, singleUse_range):
	"""
	Finds disposition of column "Rate" on the page and returns it's column-letter
	"""
	page = book.get_sheet_by_name(sheet)  
	data_beginning_row = singleUse_range[0]
	headline_row = str(data_beginning_row - 1)
	
	for cell in page[headline_row]:
		if cell.value == 'Rate':
			return get_column_letter(cell.column)
	raise ValueError(f'could not find disposition of column "Rate" (section "Single use packaging") on {sheet}')


def find_quantities_multiUse(book, sheet:str, multiUse_range):
	"""
	Finds disposition of columns: Deposit, Packaging rental charge on page
	and returns their column-letters. E.g ('S', 'V').
	"""
	page = book.get_sheet_by_name(sheet)
	data_beginning_row = multiUse_range[0]
	
	if page[f'A{data_beginning_row-1}'].value == 'Date':
		headline_row = str(data_beginning_row - 1)
	elif is_gap_after_date(page[f'{data_beginning_row-1}']):
		headline_row = str(data_beginning_row - 2)
	else:
		raise ValueError(f"couldn't find headline_row of 'Multi use packaging' on {sheet}")
	
	quantity_columns = [cell.column for cell in page[headline_row] 
					if cell.value in ('Number', 'Deposit', 'Packaging\nrental charge', 'Packaging')]
	if not len(quantity_columns) == 3:
		raise ValueError(f"couldn't find one of: 'Number', 'Deposit' or 'Packaging rental charge' columns in 'Multi use packaging'-section on {sheet}")
	return tuple(get_column_letter(i) for i in quantity_columns)


def is_gap_after_date(row):
	"""
	checks is row is an empty space after 'Date'-row in additional section or not
	(some pages can contain such rows after conversion)
	Used in 'find_additional_section' function 
	"""
	for cell in row:
		if cell.value == 'rental charge':
			return True
	return False


def get_range_from_column(book, sheet:str, col_range:tuple, column_name:str):
		"""takes tuple containing (bigining-, end-) column numbers 
		from find-functions and returns values of cells in this range
		"""
		page = book.get_sheet_by_name(sheet)
		return [i[0].value for i in 
				page[f'{column_name}{col_range[0]}':f'{column_name}{col_range[1]}']]


def is_longFormat_date(date:str):
	"""cheks is format of string is like 'dd.mm.yyyy' or not"""
	if not isinstance(date, str):
		return False
	result = re.match(r'\d{2}.\d{2}.\d{4}', date)
	return not isinstance(result, type(None))


def is_rows_merged(rowData:list):
	"""cheks is string merged and looks like ['value1\nvalue2\nvalue3'] or not"""
	cell = str(rowData[0])
	if '\n' in cell:
		return True
	return False 


def split_ifMerged(values:list):
	"""[['value1\nvalue2\nvalue3']] -> ['value1', 'value2', 'value3']"""
	if len(values) == 1 and is_rows_merged(values):
		return [i for i in values[0].split('\n')]
	else:
		return values


def check_varieties_or_custumers(data_list:list, book, sheet):
	counter = 1
	for i in data_list:
		if isinstance(i, str) and i[0].isdigit() and ' ' in i:
			continue
		else:
			raise ValueError(f"'{i}' (position {counter} in column) on page '{sheet} does not look like valid variety or costumer value'")
		counter += 1

def check_codes(codes:list, book, sheet):
	for string in codes:
		string = str(string)
		if not len(string) == 3:
			raise ValueError(f'code value "{string}" on {sheet} does not look like code')
		for i in string:
			if not i.isdigit():
				raise ValueError(f'code value "{string}" on {sheet} does not look like code')


def check_numbers(numbers:list, sheet, column_name):
	row_numb=0
	for numb in numbers:
		row_numb+=1
		numb = str(numb)
		for i in numb:
			if not i.isdigit() and i != '.':
				raise ValueError(f'Error during checking Number "{numb}" value (row {row_numb}) in {column_name} on {sheet}')
	

def check_fractinalStrings(values:list, sheet, column_name):
	"""checks strings is like '1,00' or '11,00' ('dd,dd' or 'd,dd') or not"""
	row_numb=0
	for i in values:
		row_numb+=1
		i = str(i)
		res = re.match(r'\d{1,2},\d{2}', i)
		if isinstance(res, type(None)):
			raise ValueError(f'Error during checking value "{i}" - row {row_numb} in {column_name} on {sheet} ')


def correct_priece_format(price:str, book, sheet:str):
	if not isinstance(price, int):
		raise ValueError(f'error while convert to right format value {price} from column Prise on page "{sheet}"')
	price = str(price)
	if len(price) == 3:
		return float(f'0.{price}')
	elif len(price) == 2:
		return float(f'0.0{price}')
	elif len(price) > 3:
		return float(f'{price[:-3]}.{price[-3:]}')
	else:
		raise ValueError(f'error while convert to right format value {price} from column Prise on page {sheet}')


def correct_totals_format(totals:list):
	"""finds 'total' values like 1,600 that retrieved as floats (1,600 -> 1.6) 
	and corrects them to format -> 1600, int"""
	result = []
	for i in totals:
		if isinstance(i, float):
			i = str(i).replace('.', '')
			i = '{0:0<4}'.format(i)
			result.append(int(i))
		else:
			result.append(i)
	return result


def adopt_float_format(rate_value):
	"""makes rate record suitable to convert to float 
	and converts them"""
	result = re.sub(',', '.', rate_value)
	
	while result.count('.') > 1:  # if result looks like (1.000.00), remove all excess commas
		result = re.sub(r'[.]', '', result, count=1)

	return float(result)


# retrieve_...() functions used in parse()' function

def retrieve_variety(book, sheet, data_range):
	retrieved = get_range_from_column(book, sheet, data_range, 'C')
	check_varieties_or_custumers(retrieved, book, sheet)
	return(retrieved)


def retrieve_custumer(book, sheet, data_range):
	retrieved = get_range_from_column(book, sheet, data_range, 'F')
	check_varieties_or_custumers(retrieved, book, sheet)
	return(retrieved)


def retrieve_number(book, sheet, data_range, column):
	retrieved = get_range_from_column(book, sheet, data_range, column)
	check_numbers(retrieved, sheet, 'number')
	return retrieved


def retrieve_piece(book, sheet, data_range, column):
	retrieved = get_range_from_column(book, sheet, data_range, column)
	check_numbers(retrieved, sheet, 'piece')
	return retrieved


def retrieve_total(book, sheet, data_range, column):
	retrieved = get_range_from_column(book, sheet, data_range, column)
	check_numbers(retrieved, sheet, 'total')
	return correct_totals_format(retrieved)


def retrieve_price(book, sheet, data_range, column):
	retrieved = get_range_from_column(book, sheet, data_range, column)
	return [correct_priece_format(i, book, sheet) for i in retrieved]


def retrieve_amount(book, sheet, data_range, column):
	retrieved = get_range_from_column(book, sheet, data_range, column)
	return [adopt_float_format(i) for i in retrieved]


def retrieve_code(book, sheet, data_range, column):
	retrieved = get_range_from_column(book, sheet, data_range, column)
	check_codes(retrieved, book, sheet)
	return retrieved


def retrieve_code_singleUse(book, sheet, data_range):
	retrieved = get_range_from_column(book, sheet, data_range, 'B')
	retrieved = [int(i) for i in split_ifMerged(retrieved)]
	check_codes(retrieved, book, sheet)
	return retrieved


def retrieve_rate_singleUse(book, sheet, data_range, column):
	retrieved = get_range_from_column(book, sheet, data_range, column)
	retrieved = [i for i in split_ifMerged(retrieved)]
	check_fractinalStrings(retrieved, sheet, 'rate (Single use)')
	return [adopt_float_format(i) for i in retrieved]


def retrieve_code_multiUse(book, sheet, data_range):
	retrieved = get_range_from_column(book, sheet, data_range, 'B')
	retrieved = [int(i) for i in split_ifMerged(retrieved)]
	check_codes(retrieved, book, sheet)
	return retrieved


def retrieve_deposit_multiUse(book, sheet, data_range, column):
	retrieved = get_range_from_column(book, sheet, data_range, column)
	retrieved = [i for i in retrieved if i is not None]
	retrieved = split_ifMerged(retrieved)
	check_fractinalStrings(retrieved, sheet, 'Deposit (Multi use)')
	return [adopt_float_format(i) for i in retrieved]


def retrieve_rent_multiUse(book, sheet, data_range, column):
	retrieved = get_range_from_column(book, sheet, data_range, column)
	retrieved = [i for i in retrieved if i is not None]
	retrieved = split_ifMerged(retrieved)
	check_fractinalStrings(retrieved, sheet, 'Packaging rental charge (Multi use)')
	return [adopt_float_format(i) for i in retrieved]


def parse(file):
	"""
		In 'for' loop runs through each page in document and retrieves 
		all requered columns on the page and merges list of values of each retrieved column 
		to general list of the such values. 
		E.g., 'varieties += variety' - "variety" is list of variety-values 
		from particular sheet in the document whereas "varieties" is list of all variety colums from whole document.
	"""
	book = load_workbook(file)
	sheets = book.get_sheet_names()[1:]  #list of sheets without title-page
	

	varieties = []; custumers = []; numbers = []; pieces = []
	totals = []   ; prices = []   ; amounts = []; codes = []
	
	codes_singleUse = [];  rates_singleUse = []
	codes_multiUse = [];   deposits_multiUse = []; rents_multiUse = []
	

	for sh in sheets:
		mainData_range = find_main_data(book, sh)
		if mainData_range is not None:
			quantity_colums = find_quantity_columns(book, sh, mainData_range)
			
			variety = retrieve_variety(book, sh, mainData_range)
			varieties += variety
			custumer = retrieve_custumer(book, sh, mainData_range)		
			custumers += custumer
			number = retrieve_number(book, sh, mainData_range, quantity_colums[0])
			numbers += number
			piece = retrieve_piece(book, sh, mainData_range, quantity_colums[1])
			pieces += piece
			total = retrieve_total(book, sh, mainData_range, quantity_colums[2])
			totals += total
			price = retrieve_price(book, sh, mainData_range, quantity_colums[3])
			prices += price
			amount = retrieve_amount(book, sh, mainData_range, quantity_colums[4])
			amounts += amount
			code = retrieve_code(book, sh, mainData_range, quantity_colums[5])
			codes += code
		
		singleUse_range = find_additional_section(book, sh, 'Single use packaging')
		if singleUse_range is not None:
			rate_disposition = find_rates_singleUse(book, sh, singleUse_range)
			
			code_singleUse = retrieve_code_singleUse(book, sh, singleUse_range)
			codes_singleUse += code_singleUse
			rate_singleUse = retrieve_rate_singleUse(book, sh, singleUse_range, rate_disposition)
			rates_singleUse += rate_singleUse
		
		multiUse_range = find_additional_section(book, sh, 'Multi use packaging')
		if multiUse_range is not None:
			quantities_multiUse = find_quantities_multiUse(book, sh, multiUse_range)
			
			code_multiUse = retrieve_code_multiUse(book, sh, multiUse_range)
			codes_multiUse += code_multiUse
			deposit_multiUse = retrieve_deposit_multiUse(book, sh, multiUse_range, quantities_multiUse[1])
			deposits_multiUse += deposit_multiUse
			rent_multiUse = retrieve_rent_multiUse(book, sh, multiUse_range, quantities_multiUse[2]) 
			rents_multiUse += rent_multiUse


	return Fieldstuple(varieties, custumers, numbers, pieces, 
			   totals,    prices,    amounts, codes, 
			   codes_singleUse, rates_singleUse,
			   codes_multiUse,  deposits_multiUse, rents_multiUse) 


if __name__ == '__main__':
	pass
